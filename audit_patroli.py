import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError, ImageFilter
import imagehash
import io
import os
import re
import sqlite3
import zipfile
import hashlib
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import math

# =========================================================
# CONFIG UI
# =========================================================
st.set_page_config(page_title="Audit Foto Patroli", layout="wide")
st.title("🕵️ AUDIT FOTO PATROLI")
st.caption(
    "Audit foto patroli duplicate (Embedded Excel + Google Docs/Drive). "
    "✅ Yang DIAUDIT (VALID/GUGUR/CEK MANUAL) cuma foto patroli yang ada overlay GEO (kotak gelap + teks putih timestamp/koordinat biasanya di bawah). "
    "✅ Foto GEO boleh ada logo perusahaan di pojok (tetap diaudit). "
    "✅ Foto 'logo doang' (hitam/putih, banner, template) -> SKIP (Logo-Only), nggak masuk audit. "
    "✅ Hash duplikat pakai MID+GEO crop (logo atas tidak ikut), jadi logo vendor tidak bikin semua foto dianggap duplikat. "
    "✅ Output audit jelasin 'DUP OF' (rujukan duplikat: file/sheet/lokasi/segment/url) untuk duplikat internal & lintas bulan (DB)."
)

# =========================================================
# STREAMLIT UI (UPLOADER HARUS DI ATAS)
# =========================================================
uploaded = st.file_uploader("Upload Excel Patroli (.xlsx)", type=["xlsx"])

colA, colB = st.columns([1, 2])
with colA:
    preview_limit = st.number_input("Maks preview gambar", min_value=0, max_value=500, value=200, step=10)
with colB:
    st.info("Reset history ada di sidebar kiri. Logo-only tidak ikut audit. Hash audit pakai MID+GEO crop supaya logo vendor tidak bikin semua foto dianggap duplikat.")

# =========================================================
# DATABASE
# =========================================================
DB_PATH = "audit_history.db"

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS history (
            sha256 TEXT PRIMARY KEY,
            phash  TEXT,
            source_type TEXT,
            source_file TEXT,
            sheet TEXT,
            location TEXT,
            cluster TEXT,
            segment TEXT,
            url TEXT,
            first_seen DATE
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_phash ON history(phash)")
    return conn

def db_lookup(conn, sha256_hex: str, phash_str: str):
    exact = conn.execute(
        "SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE sha256=?",
        (sha256_hex,)
    ).fetchone()

    ph = conn.execute(
        "SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE phash=? LIMIT 1",
        (phash_str,)
    ).fetchone()

    return exact, ph

def db_insert(conn, row: dict):
    conn.execute("""
        INSERT OR IGNORE INTO history
        (sha256, phash, source_type, source_file, sheet, location, cluster, segment, url, first_seen)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row["sha256"], row["phash"], row["source_type"], row["source_file"],
        row["sheet"], row["location"], row["cluster"], row["segment"], row["url"], row["first_seen"]
    ))

# =========================================================
# RESET / HAPUS HISTORY AUDIT (SIDEBAR)
# =========================================================
st.sidebar.subheader("🧹 Reset History Audit")
confirm_reset = st.sidebar.checkbox("Saya yakin mau hapus total history", value=False)

if st.sidebar.button("🗑️ HAPUS TOTAL HISTORY (RESET)"):
    if not confirm_reset:
        st.sidebar.warning("Centang konfirmasi dulu.")
    else:
        try:
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
                st.sidebar.success("✅ History dihapus. App akan refresh.")
                st.rerun()
            else:
                st.sidebar.info("ℹ️ History sudah kosong (file DB tidak ada).")
        except Exception as e:
            st.sidebar.error(f"❌ Gagal hapus DB: {e}")

# =========================================================
# HASHING
# =========================================================
def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

# =========================================================
# METRIK VISUAL (LOGO-ONLY)
# =========================================================
def image_entropy_gray(img: Image.Image) -> float:
    g = img.convert("L").resize((256, 256))
    hist = g.histogram()
    total = sum(hist)
    if total == 0:
        return 0.0
    ent = 0.0
    for h in hist:
        if h:
            p = h / total
            ent -= p * math.log2(p)
    return ent

def edge_density(img: Image.Image) -> float:
    g = img.convert("L").resize((256, 256)).filter(ImageFilter.FIND_EDGES)
    px = list(g.getdata())
    mean = sum(px) / (len(px) or 1)
    return mean / 255.0

def unique_color_ratio(img: Image.Image, k=24) -> float:
    small = img.resize((256, 256))
    q = small.quantize(colors=k, method=2)  # adaptive
    px = list(q.getdata())
    uniq = len(set(px))
    return uniq / float(k)

def bright_ratio(img: Image.Image, thr=240) -> float:
    g = img.convert("L").resize((256, 256))
    px = list(g.getdata())
    return sum(1 for p in px if p >= thr) / (len(px) or 1)

def is_logo_only(img: Image.Image) -> tuple[bool, str]:
    """
    Target: logo/banner/template yang cuma logo doang (hitam/putih).
    HARUS SKIP.

    Heuristik:
      - edge rendah (flat)
      - warna sedikit
      - salah satu:
        a) background putih dominan (logo putih)
        b) entropi rendah (logo gelap / flat)
    """
    w, h = img.size
    if w < 140 or h < 140:
        return True, "LogoOnly: terlalu kecil"

    ent = image_entropy_gray(img)
    ed  = edge_density(img)
    ucr = unique_color_ratio(img, k=24)
    br  = bright_ratio(img, thr=240)

    # logo putih: putih dominan + warna sedikit + edge rendah
    if br >= 0.58 and ucr <= 0.45 and ed <= 0.070:
        return True, f"LogoOnly(WhiteBG): bright={br:.2f}, ucol={ucr:.2f}, edge={ed:.3f}"

    # logo gelap/flat: entropi rendah + warna sedikit + edge rendah
    if ent <= 4.20 and ucr <= 0.45 and ed <= 0.070:
        return True, f"LogoOnly(Flat): ent={ent:.2f}, ucol={ucr:.2f}, edge={ed:.3f}"

    return False, f"NotLogo: ent={ent:.2f}, ucol={ucr:.2f}, edge={ed:.3f}"

# =========================================================
# DETEKSI GEO OVERLAY (SCAN BAWAH)
# =========================================================
def _patch_stats_gray(patch_gray: Image.Image):
    px = list(patch_gray.getdata())
    n = len(px) or 1
    mean = sum(px) / n
    var = sum((p - mean) ** 2 for p in px) / n
    std = math.sqrt(var)
    white_ratio = sum(1 for p in px if p >= 220) / n
    dark_ratio  = sum(1 for p in px if p <= 70) / n
    # edge patch: teks biasanya bikin edge naik
    pedge = edge_density(patch_gray.convert("RGB"))
    return mean, std, white_ratio, dark_ratio, pedge

def overlay_geo_best(img: Image.Image) -> tuple[bool, str, tuple[int,int,int,int] | None]:
    """
    Scan 3 area bawah: kiri, tengah, kanan.
    Output:
      - ok (bool)
      - dbg string
      - roi bbox terbaik (X0,Y0,X1,Y1) untuk crop GEO
    """
    w, h = img.size
    if w < 240 or h < 240:
        return False, "NonGEO: ukuran kecil", None

    rois = [
        ("LB", (0.00, 0.62, 0.50, 1.00)),
        ("MB", (0.20, 0.62, 0.80, 1.00)),
        ("RB", (0.50, 0.62, 1.00, 1.00)),
    ]

    best_score = -1
    best_dbg = ""
    best_bbox = None

    for name, (x0, y0, x1, y1) in rois:
        X0 = int(w * x0); Y0 = int(h * y0)
        X1 = int(w * x1); Y1 = int(h * y1)

        patch = img.crop((X0, Y0, X1, Y1)).convert("L").resize((260, 260))
        mean, std, white_ratio, dark_ratio, pedge = _patch_stats_gray(patch)

        has_white_text = white_ratio >= 0.003
        has_dark_box   = dark_ratio  >= 0.006
        has_contrast   = std >= 14
        has_patch_edge = pedge >= 0.030

        score = int(has_white_text) + int(has_dark_box) + int(has_contrast) + int(has_patch_edge)
        dbg = (
            f"{name}: mean={mean:.1f}, std={std:.1f}, "
            f"white={white_ratio:.3f}, dark={dark_ratio:.3f}, edge={pedge:.3f}, score={score}"
        )

        if score > best_score:
            best_score = score
            best_dbg = dbg
            best_bbox = (X0, Y0, X1, Y1)

    if best_score >= 2:
        return True, f"GEO✅ ({best_dbg})", best_bbox
    return False, f"NonGEO ({best_dbg})", best_bbox

# =========================================================
# MID+GEO CROP untuk PHASH (logo atas tidak ikut)
# =========================================================
def crop_mid(img: Image.Image) -> Image.Image:
    w, h = img.size
    x0 = int(w * 0.10)
    x1 = int(w * 0.90)
    y0 = int(h * 0.15)   # buang atas (logo)
    y1 = int(h * 0.75)   # jangan kebanyakan bawah
    if x1 <= x0 or y1 <= y0:
        return img
    return img.crop((x0, y0, x1, y1))

def crop_geo_best_patch(img: Image.Image, bbox: tuple[int,int,int,int] | None) -> Image.Image:
    w, h = img.size
    if bbox:
        X0, Y0, X1, Y1 = bbox
        if X1 > X0 and Y1 > Y0:
            return img.crop((X0, Y0, X1, Y1))
    # fallback RB
    return img.crop((int(w*0.50), int(h*0.62), w, h))

def compose_audit_image(img: Image.Image, geo_bbox: tuple[int,int,int,int] | None) -> Image.Image:
    """
    Gabung MID + GEO menjadi satu canvas -> dipakai untuk phash.
    """
    mid = crop_mid(img).resize((512, 512))
    geo = crop_geo_best_patch(img, geo_bbox).resize((512, 256))

    canvas = Image.new("RGB", (512, 768))
    canvas.paste(mid, (0, 0))
    canvas.paste(geo, (0, 512))
    return canvas

def compute_hashes_from_bytes(img_bytes: bytes):
    """
    Return: sha256(full bytes), phash(MID+GEO), thumb(preview), full_img
    """
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except UnidentifiedImageError:
        return None, None, None, None

    thumb = img.copy()
    thumb.thumbnail((260, 260))

    sh = sha256_bytes(img_bytes)

    # geo bbox untuk hashing (kalau bukan geo, tetap bikin bbox terbaik agar stabil)
    _, _, geo_bbox = overlay_geo_best(img)
    audit_img = compose_audit_image(img, geo_bbox)
    ph = str(imagehash.phash(audit_img))

    return sh, ph, thumb, img

# =========================================================
# CLASSIFY FINAL
# =========================================================
def classify_for_audit(img: Image.Image) -> tuple[bool, str, str]:
    """
    RULE FINAL:
    1) Kalau ada GEO overlay -> MASUK AUDIT (logo pojok bebas)
    2) Kalau tidak GEO:
         - kalau logo-only -> SKIP (Logo-Only)
         - selain itu -> SKIP (Non-Patroli)
    """
    ok_geo, geo_dbg, _ = overlay_geo_best(img)
    if ok_geo:
        return True, "", ""

    logo, logo_dbg = is_logo_only(img)
    if logo:
        return False, "⏭️ SKIP (Logo-Only)", logo_dbg

    return False, "⏭️ SKIP (Non-Patroli)", geo_dbg

# =========================================================
# GOOGLE LINK HANDLING
# =========================================================
DOC_ID_RE = re.compile(r"/document/d/([a-zA-Z0-9_-]+)")
DRIVE_FILE_ID_RE = re.compile(r"/file/d/([a-zA-Z0-9_-]+)")
GENERIC_ID_RE = re.compile(r"(?:id=)([a-zA-Z0-9_-]+)")

def build_gdocs_export_docx_url(doc_id: str) -> str:
    return f"https://docs.google.com/document/d/{doc_id}/export?format=docx"

def build_drive_download_url(file_id: str) -> str:
    return f"https://drive.google.com/uc?export=download&id={file_id}"

def http_get_bytes(url: str, timeout=25):
    try:
        r = requests.get(url, timeout=timeout, stream=True, allow_redirects=True, headers={
            "User-Agent": "PatrolPhotoAudit/1.0"
        })
        if r.status_code != 200:
            return None
        return r.content
    except requests.RequestException:
        return None

def extract_images_from_docx_bytes(docx_bytes: bytes) -> list[bytes]:
    out = []
    with zipfile.ZipFile(io.BytesIO(docx_bytes), "r") as z:
        for name in z.namelist():
            if name.startswith("word/media/"):
                try:
                    out.append(z.read(name))
                except:
                    pass
    return out

def download_images_from_url(url: str) -> list[bytes]:
    if not url or not isinstance(url, str):
        return []

    m = DOC_ID_RE.search(url)
    if m:
        doc_id = m.group(1)
        docx_bytes = http_get_bytes(build_gdocs_export_docx_url(doc_id))
        if not docx_bytes:
            return []
        return extract_images_from_docx_bytes(docx_bytes)

    m = DRIVE_FILE_ID_RE.search(url)
    if m:
        b = http_get_bytes(build_drive_download_url(m.group(1)))
        return [b] if b else []

    m = GENERIC_ID_RE.search(url)
    if m and "google" in url:
        b = http_get_bytes(build_drive_download_url(m.group(1)))
        return [b] if b else []

    return []

# =========================================================
# EXCEL PARSING
# =========================================================
def find_header_row_and_cols(ws, max_scan_rows=40):
    target = {"cluster": ["cluster"], "segment": ["segment", "segment name", "segmen"], "link": ["link", "url"]}
    def norm(v): return str(v).strip().lower() if v is not None else ""

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        if not any(row_vals):
            continue

        col_cluster = col_segment = col_link = None
        for idx, val in enumerate(row_vals, start=1):
            if any(k in val for k in target["cluster"]): col_cluster = idx
            if any(k in val for k in target["segment"]): col_segment = idx
            if any(k in val for k in target["link"]): col_link = idx

        if col_cluster and col_segment and col_link:
            return r, col_cluster, col_segment, col_link

    return 4, 2, 3, 7

def extract_embedded_images(wb, source_file_name: str):
    items = []
    for ws in wb.worksheets:
        imgs = getattr(ws, "_images", [])
        if not imgs:
            continue

        header_row, col_cluster, col_segment, _ = find_header_row_and_cols(ws)

        for img_obj in imgs:
            try:
                row = img_obj.anchor._from.row + 1
                col = img_obj.anchor._from.col + 1
                cluster = ws.cell(row=row, column=col_cluster).value or "N/A"
                segment = ws.cell(row=row, column=col_segment).value or "N/A"

                raw = img_obj._data()
                sh, ph, thumb, full_img = compute_hashes_from_bytes(raw)
                if full_img is None:
                    continue

                audit_ok, skip_status, skip_reason = classify_for_audit(full_img)
                if not audit_ok:
                    items.append({
                        "source_type": "EmbeddedExcel",
                        "source_file": source_file_name,
                        "sheet": ws.title,
                        "location": f"R{row}C{col}",
                        "cluster": str(cluster),
                        "segment": str(segment),
                        "url": "",
                        "sha256": "",
                        "phash": "",
                        "status_akhir": skip_status,
                        "skip_reason": skip_reason or "",
                        "thumb": thumb
                    })
                    continue

                items.append({
                    "source_type": "EmbeddedExcel",
                    "source_file": source_file_name,
                    "sheet": ws.title,
                    "location": f"R{row}C{col}",
                    "cluster": str(cluster),
                    "segment": str(segment),
                    "url": "",
                    "sha256": sh,
                    "phash": ph,
                    "status_akhir": "",
                    "skip_reason": "",
                    "thumb": thumb
                })
            except:
                continue
    return items

def extract_link_images(wb, source_file_name: str, max_workers=12):
    jobs = []
    for ws in wb.worksheets:
        header_row, col_cluster, col_segment, col_link = find_header_row_and_cols(ws)
        for r in range(header_row + 1, ws.max_row + 1):
            url = ws.cell(r, col_link).value
            if not url:
                continue
            url = str(url)
            if "google" not in url:
                continue

            cluster = ws.cell(r, col_cluster).value or "N/A"
            segment = ws.cell(r, col_segment).value or "N/A"
            jobs.append((ws.title, r, col_link, str(cluster), str(segment), url))

    items = []
    if not jobs:
        return items

    def worker(job):
        sheet, r, col_link, cluster, segment, url = job
        img_bytes_list = download_images_from_url(url)

        out = []
        for idx, b in enumerate(img_bytes_list, start=1):
            if not b:
                continue
            sh, ph, thumb, full_img = compute_hashes_from_bytes(b)
            if full_img is None:
                continue

            audit_ok, skip_status, skip_reason = classify_for_audit(full_img)
            if not audit_ok:
                out.append({
                    "source_type": "CloudLink",
                    "source_file": source_file_name,
                    "sheet": sheet,
                    "location": f"R{r}C{col_link}#IMG{idx}",
                    "cluster": cluster,
                    "segment": segment,
                    "url": url,
                    "sha256": "",
                    "phash": "",
                    "status_akhir": skip_status,
                    "skip_reason": skip_reason or "",
                    "thumb": thumb
                })
                continue

            out.append({
                "source_type": "CloudLink",
                "source_file": source_file_name,
                "sheet": sheet,
                "location": f"R{r}C{col_link}#IMG{idx}",
                "cluster": cluster,
                "segment": segment,
                "url": url,
                "sha256": sh,
                "phash": ph,
                "status_akhir": "",
                "skip_reason": "",
                "thumb": thumb
            })
        return out

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(worker, j) for j in jobs]
        for f in as_completed(futures):
            try:
                items.extend(f.result())
            except:
                pass

    return items

# =========================================================
# DUP OF (INTERNAL + HISTORY)
# =========================================================
def _ref_string(row: dict) -> str:
    return (
        f'{row.get("source_file","")} | {row.get("sheet","")} | {row.get("location","")} | '
        f'seg={row.get("segment","")} | url={row.get("url","")}'
    )

def apply_dup_of_columns(df_audit: pd.DataFrame, conn) -> pd.DataFrame:
    df = df_audit.copy()
    for c in ["dup_of_type", "dup_of_detail"]:
        if c not in df.columns:
            df[c] = ""

    # INTERNAL EXACT: sha256
    first_by_sha = {}
    for i, r in df.iterrows():
        sha = str(r.get("sha256", "")).strip()
        if not sha:
            continue
        if sha not in first_by_sha:
            first_by_sha[sha] = i
        else:
            ref = df.loc[first_by_sha[sha]].to_dict()
            df.at[i, "dup_of_type"] = "INTERNAL_EXACT"
            df.at[i, "dup_of_detail"] = _ref_string(ref)

    # INTERNAL PHASH: (hash MID+GEO)
    first_by_ph = {}
    for i, r in df.iterrows():
        if str(df.at[i, "dup_of_type"]).strip():
            continue
        ph = str(r.get("phash", "")).strip()
        if not ph:
            continue
        if ph not in first_by_ph:
            first_by_ph[ph] = i
        else:
            ref = df.loc[first_by_ph[ph]].to_dict()
            df.at[i, "dup_of_type"] = "INTERNAL_PHASH"
            df.at[i, "dup_of_detail"] = _ref_string(ref)

    # HISTORY (kalau belum kena internal)
    for i, r in df.iterrows():
        if str(df.at[i, "dup_of_type"]).strip():
            continue
        sha = str(r.get("sha256", "")).strip()
        ph  = str(r.get("phash", "")).strip()
        exact, sim = db_lookup(conn, sha, ph)

        if exact:
            df.at[i, "dup_of_type"] = "HISTORY_EXACT"
            df.at[i, "dup_of_detail"] = (
                f'{exact[0]} | {exact[1]} | {exact[2]} | seg={exact[4]} | url={exact[5]} | first={exact[6]}'
            )
        elif sim:
            df.at[i, "dup_of_type"] = "HISTORY_PHASH"
            df.at[i, "dup_of_detail"] = (
                f'{sim[0]} | {sim[1]} | {sim[2]} | seg={sim[4]} | url={sim[5]} | first={sim[6]}'
            )

    return df

# =========================================================
# AUDIT LOGIC
# =========================================================
def audit_workbook(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    source_file_name = os.path.basename(xlsx_path)

    df = pd.DataFrame(extract_embedded_images(wb, source_file_name) + extract_link_images(wb, source_file_name))
    if df.empty:
        return df

    audited_mask = (df["status_akhir"].astype(str).str.strip() == "") & (df["sha256"].astype(str).str.strip() != "")
    df_audit = df[audited_mask].copy()
    df_skip = df[~audited_mask].copy()

    base_cols = [
        "dup_internal_exact","dup_internal_phash","history_status","history_detail","first_seen",
        "dup_of_type","dup_of_detail"
    ]

    if df_audit.empty:
        for col in base_cols:
            df[col] = ""
        return df

    # duplikat internal (exact bytes & phash MID+GEO)
    df_audit["dup_internal_exact"] = df_audit.duplicated("sha256", keep="first")
    df_audit["dup_internal_phash"] = df_audit.duplicated("phash", keep="first")

    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    df_audit["first_seen"] = today

    # DUP OF detail
    df_audit = apply_dup_of_columns(df_audit, conn)

    # history status (human-friendly)
    hist_status, hist_detail = [], []
    for _, row in df_audit.iterrows():
        exact, ph = db_lookup(conn, row["sha256"], row["phash"])
        if exact:
            hist_status.append("REUPLOAD_EXACT")
            hist_detail.append(f"Pernah terbit {exact[-1]} | {exact[0]} | {exact[1]} | {exact[2]}")
        elif ph:
            hist_status.append("REUPLOAD_SIMILAR_PHASH")
            hist_detail.append(f"Mirip phash: {ph[-1]} | {ph[0]} | {ph[1]} | {ph[2]}")
        else:
            hist_status.append("NEW")
            hist_detail.append("")
    df_audit["history_status"] = hist_status
    df_audit["history_detail"] = hist_detail

    def decide(r):
        if r["history_status"] == "REUPLOAD_EXACT":
            return "❌ GUGUR (Pernah Terbit - Exact)"
        if r["history_status"] == "REUPLOAD_SIMILAR_PHASH":
            return "⚠️ CEK MANUAL (Mirip Foto Lama)"
        if r["dup_internal_exact"]:
            return "❌ GUGUR (Duplikat di File Ini - Exact)"
        if r["dup_internal_phash"]:
            return "⚠️ CEK MANUAL (Duplikat Mirip di File Ini)"
        return "✅ VALID"

    df_audit["status_akhir"] = df_audit.apply(decide, axis=1)

    # simpan hanya VALID ke history (lintas bulan)
    for _, r in df_audit[df_audit["status_akhir"] == "✅ VALID"].iterrows():
        db_insert(conn, {
            "sha256": r["sha256"],
            "phash": r["phash"],
            "source_type": r["source_type"],
            "source_file": r["source_file"],
            "sheet": r["sheet"],
            "location": r["location"],
            "cluster": r["cluster"],
            "segment": r["segment"],
            "url": r["url"],
            "first_seen": r["first_seen"]
        })
    conn.commit()
    conn.close()

    # rapihin skip col
    for col in base_cols:
        if col not in df_skip.columns:
            df_skip[col] = ""

    return pd.concat([df_audit, df_skip], ignore_index=True)

# =========================================================
# RUN UI ACTIONS
# =========================================================
if uploaded:
    tmp_path = "temp_upload.xlsx"
    with open(tmp_path, "wb") as f:
        f.write(uploaded.getbuffer())

    if st.button("🚀 MULAI AUDIT"):
        with st.status("Meng-audit foto…", expanded=True) as status:
            df = audit_workbook(tmp_path)

            if df.empty:
                status.update(label="Tidak ada foto yang terbaca.", state="error")
                st.warning("Tidak ditemukan foto embedded atau foto dari link Google yang bisa diproses.")
            else:
                status.update(label="Audit selesai.", state="complete")

                audited_only = df[~df["status_akhir"].astype(str).str.startswith("⏭️ SKIP")].copy()

                st.subheader("Ringkasan (HANYA yang diaudit)")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Total Diaudit", len(audited_only))
                c2.metric("VALID", int((audited_only["status_akhir"] == "✅ VALID").sum()))
                c3.metric("GUGUR", int(audited_only["status_akhir"].astype(str).str.contains("GUGUR").sum()))
                c4.metric("CEK MANUAL", int(audited_only["status_akhir"].astype(str).str.contains("CEK MANUAL").sum()))

                st.subheader("Laporan (Termasuk SKIP)")
                report = df.drop(columns=["thumb"], errors="ignore")
                st.dataframe(report, use_container_width=True)

                outbuf = io.BytesIO()
                report.to_excel(outbuf, index=False)
                today = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    "📥 Download Laporan (Excel)",
                    data=outbuf.getvalue(),
                    file_name=f"Laporan_Audit_Foto_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("Preview (dibatasi)")
                shown = 0
                cols = st.columns(4)

                def prio(s):
                    s = str(s)
                    if "GUGUR" in s or "CEK MANUAL" in s:
                        return 0
                    if s.startswith("✅"):
                        return 1
                    if s.startswith("⏭️ SKIP"):
                        return 2
                    return 3

                df_view = df.copy()
                df_view["_p"] = df_view["status_akhir"].map(prio)
                df_view = df_view.sort_values("_p").drop(columns=["_p"])

                for _, r in df_view.iterrows():
                    if preview_limit == 0 or shown >= preview_limit:
                        break
                    if r.get("thumb") is None:
                        continue
                    with cols[shown % 4]:
                        st.image(
                            r["thumb"],
                            caption=f'{r.get("sheet","")} | {r.get("location","")}\n{r.get("status_akhir","")}'
                        )
                        if r.get("skip_reason"):
                            st.caption(str(r["skip_reason"]))
                        if r.get("dup_of_type"):
                            st.caption(f'DUP OF [{r["dup_of_type"]}]: {r.get("dup_of_detail","")}')
                        if r.get("history_detail"):
                            st.caption(str(r["history_detail"]))
                    shown += 1

    if os.path.exists(tmp_path):
        os.remove(tmp_path)

st.divider()
st.caption(f"DB history disimpan lokal: {DB_PATH} (jangan dihapus kalau mau deteksi lintas bulan)")
