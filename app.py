import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import imagehash
import io
import os
import sqlite3
import requests

# --- KONFIGURASI ---
st.set_page_config(page_title="Audit Foto Patroli Multi-Source", layout="wide")
st.title("🔍 Audit Foto Patroli Fiber Optic (Bulan 1-4+)")
st.write("Mendukung foto langsung di Excel & Link Google Drive.")

# --- DATABASE (Ingatan Lintas Bulan) ---
def get_db_connection():
    conn = sqlite3.connect('audit_history.db')
    conn.execute('''CREATE TABLE IF NOT EXISTS foto_history 
                 (hash TEXT PRIMARY KEY, cluster TEXT, segment TEXT, tanggal TEXT)''')
    return conn

def cek_history(p_hash):
    with get_db_connection() as conn:
        return conn.execute("SELECT cluster, tanggal FROM foto_history WHERE hash=?", (p_hash,)).fetchone()

def simpan_history(df):
    with get_db_connection() as conn:
        for _, r in df.iterrows():
            try:
                conn.execute("INSERT OR IGNORE INTO foto_history VALUES (?,?,?,?)", 
                             (r['Hash'], r['Cluster'], r['Segment Name'], r['Tanggal Patroli']))
            except: pass

# --- FUNGSI DOWNLOAD GOOGLE DRIVE ---
def download_gdrive_img(url):
    try:
        file_id = ""
        if 'id=' in url:
            file_id = url.split('id=')[-1].split('&')[0]
        elif '/d/' in url:
            file_id = url.split('/d/')[1].split('/')[0]
        
        if file_id:
            direct_url = f'https://drive.google.com/uc?export=download&id={file_id}'
            response = requests.get(direct_url, timeout=5) # Timeout dipersingkat agar tidak hang
            img = Image.open(io.BytesIO(response.content))
            img.thumbnail((400, 400)) # OPTIMASI: Kecilkan ukuran di memori
            return img
    except:
        return None
    return None

# --- PROSES EXCEL ---
def proses_excel(file):
    wb = load_workbook(file, data_only=True)
    results = []
    
    for sheet in wb.worksheets:
        # 1. AMBIL FOTO TERTANAM
        if hasattr(sheet, '_images'):
            st.info(f"Mendeteksi {len(sheet._images)} gambar tertanam di sheet: {sheet.title}")
            for img_obj in sheet._images:
                try:
                    row = img_obj.anchor._from.row + 1
                    col = img_obj.anchor._from.col + 1
                    
                    c_val = sheet.cell(row=row, column=1).value
                    s_val = sheet.cell(row=row, column=2).value
                    t_val = sheet.cell(row=row, column=3).value
                    
                    img = Image.open(io.BytesIO(img_obj._data()))
                    img.thumbnail((400, 400)) # OPTIMASI RAM
                    h = str(imagehash.phash(img))
                    hist = cek_history(h)
                    
                    results.append({
                        "Cluster": str(c_val) if c_val else "N/A",
                        "Segment Name": str(s_val) if s_val else "N/A",
                        "Tanggal Patroli": str(t_val) if t_val else "N/A",
                        "Posisi": f"Baris {row}, Kol {get_column_letter(col)}",
                        "Tipe": "Embedded Image",
                        "Hash": h,
                        "Hist_Info": f"⚠️ Dulu di {hist[0]} ({hist[1]})" if hist else "✅ NEW",
                        "Img_Obj": img
                    })
                except: continue

        # 2. AMBIL FOTO DARI LINK GDRIVE (Optimasi Safe Loading)
        # Cari baris yang ada link GDrive saja
        max_r = min(sheet.max_row, 500) # Batasi 500 baris per sheet agar tidak overload
        for r_idx in range(1, max_r + 1):
            cell_url = sheet.cell(row=r_idx, column=5).value
            if cell_url and isinstance(cell_url, str) and "drive.google.com" in cell_url:
                with st.status(f"Mendownload baris {r_idx}...", expanded=False) as status:
                    img = download_gdrive_img(cell_url)
                    if img:
                        h = str(imagehash.phash(img))
                        hist = cek_history(h)
                        results.append({
                            "Cluster": str(sheet.cell(row=r_idx, column=1).value) or "N/A",
                            "Segment Name": str(sheet.cell(row=r_idx, column=2).value) or "N/A",
                            "Tanggal Patroli": str(sheet.cell(row=r_idx, column=3).value) or "N/A",
                            "Posisi": f"Baris {r_idx}, Kol E (Link)",
                            "Tipe": "GDrive Link",
                            "Hash": h,
                            "Hist_Info": f"⚠️ Dulu di {hist[0]} ({hist[1]})" if hist else "✅ NEW",
                            "Img_Obj": img
                        })
                        status.update(label=f"Baris {r_idx} Selesai!", state="complete")
                    else:
                        status.update(label=f"Baris {r_idx} Gagal Download", state="error")

    return results

# --- UI ---
up = st.file_uploader("Upload Excel Patroli (Bulan 1-4+)", type=["xlsx"])

if up:
    if st.button("🚀 Mulai Audit Data"):
        with st.spinner('Menganalisis foto & link...'):
            with open("temp.xlsx", "wb") as f: f.write(up.getbuffer())
            data = proses_excel("temp.xlsx")
        
        if data:
            df = pd.DataFrame(data)
            dup_internal = df.duplicated('Hash', keep=False)
            
            def set_status(i, r):
                if "Dulu" in r['Hist_Info']: return "❌ DUPLICATE (HISTORY)"
                if dup_internal[i]: return "❌ DUPLICATE (INTERNAL)"
                return "✅ REAL PICT"
                
            df['Status Audit'] = [set_status(i, r) for i, r in df.iterrows()]
            
            st.success(f"Analisis Selesai! {len(df)} foto diproses.")
            
            # --- ACTION BUTTONS ---
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Simpan ke Memori Audit"):
                    simpan_history(df)
                    st.success("Data berhasil diingat untuk bulan depan!")
            
            with col2:
                out = df.drop(columns=['Img_Obj'])
                towrite = io.BytesIO()
                out.to_excel(towrite, index=False)
                st.download_button("📥 Download Excel Hasil Audit", towrite.getvalue(), "Hasil_Audit_Patroli.xlsx")

            # --- DISPLAY TABS ---
            t1, t2 = st.tabs(["📊 Tabel Ringkasan", "🚩 Galeri Visual"])
            with t1:
                st.dataframe(df[['Cluster', 'Segment Name', 'Status Audit', 'Hist_Info', 'Posisi']], use_container_width=True)
            with t2:
                for _, r in df.iterrows():
                    color = "red" if "❌" in r['Status Audit'] else "green"
                    st.markdown(f"### :{color}[{r['Status Audit']}]")
                    st.write(f"**Segmen:** {r['Segment Name']} | **Posisi:** {r['Posisi']}")
                    st.image(r['Img_Obj'], width=400)
                    st.divider()

    if os.path.exists("temp.xlsx"): os.remove("temp.xlsx")
