import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import imagehash
import io
import os
import sqlite3
import requests
import re

# --- KONFIGURASI ---
st.set_page_config(page_title="Audit Foto Duplicate", layout="wide")
# Perubahan Judul sesuai permintaan
st.title("🛡️ Audit Foto Duplicate")
st.markdown("### Deteksi Otomatis Duplikasi")

# --- DATABASE ENGINE ---
def get_db_connection():
    conn = sqlite3.connect('audit_history_master.db')
    conn.execute('''CREATE TABLE IF NOT EXISTS history 
                 (hash TEXT PRIMARY KEY, info TEXT, date TEXT)''')
    return conn

def cek_global_history(p_hash):
    with get_db_connection() as conn:
        return conn.execute("SELECT info, date FROM history WHERE hash=?", (p_hash,)).fetchone()

# --- INTELLIGENT DOWNLOADER ---
def smart_download(url):
    try:
        # Regex yang lebih kuat untuk menangkap ID file dari GDocs/GDrive
        regex = r"(?<=/d/|id=)([a-zA-Z0-9-_]+)"
        match = re.search(regex, url)
        if not match: return None
        
        file_id = match.group(1)
        direct_url = f'https://drive.google.com/uc?export=download&id={file_id}'
        
        # Request dengan Timeout agar tidak hang jika koneksi lambat
        response = requests.get(direct_url, timeout=10, stream=True)
        if response.status_code == 200:
            img = Image.open(io.BytesIO(response.content))
            img = img.convert('RGB') # Standarisasi format
            img.thumbnail((300, 300)) # Optimasi RAM
            return img
    except:
        return None
    return None

# --- CORE AUDIT ENGINE ---
def run_deep_audit(file_path):
    wb = load_workbook(file_path, data_only=True)
    results = []
    
    for sheet in wb.worksheets:
        # A. PROSES FOTO YANG MENEMPEL (EMBEDDED)
        if hasattr(sheet, '_images'):
            for img_obj in sheet._images:
                try:
                    row_idx = img_obj.anchor._from.row + 1
                    col_idx = img_obj.anchor._from.col + 1
                    
                    # Metadata (Asumsi Kolom B=Cluster, C=Segment)
                    cluster = sheet.cell(row=row_idx, column=2).value or "N/A"
                    segment = sheet.cell(row=row_idx, column=3).value or "N/A"
                    
                    img = Image.open(io.BytesIO(img_obj._data())).convert('RGB')
                    h = str(imagehash.phash(img))
                    
                    results.append({
                        "Source": "Embedded",
                        "Location": f"{sheet.title} | Baris {row_idx}",
                        "Cluster": cluster, "Segment": segment,
                        "Hash": h, "Img_Obj": img
                    })
                except: continue

        # B. PROSES FOTO DARI LINK (GDRIVE/GDOCS)
        # Scan kolom LINK (Kolom G = indeks 7)
        for r in range(1, sheet.max_row + 1):
            cell_val = sheet.cell(row=r, column=7).value
            if cell_val and "google.com" in str(cell_val):
                with st.status(f"Mengecek Link Baris {r}...", expanded=False) as status:
                    img = smart_download(str(cell_val))
                    if img:
                        h = str(imagehash.phash(img))
                        results.append({
                            "Source": "Cloud Link",
                            "Location": f"Baris {r} (Kol G)",
                            "Cluster": sheet.cell(row=r, column=2).value or "N/A",
                            "Segment": sheet.cell(row=r, column=3).value or "N/A",
                            "Hash": h, "Img_Obj": img
                        })
                        status.update(label=f"Baris {r} Selesai!", state="complete")
                    else:
                        status.update(label=f"Baris {r} Gagal diunduh", state="error")
    return results

# --- UI LOGIC ---
uploaded = st.file_uploader("Upload File Excel Patroli (.xlsx)", type=["xlsx"])

if uploaded:
    if st.button("🚀 JALANKAN FULL AUDIT"):
        with st.spinner('Menganalisis data... Mohon tunggu.'):
            # Simpan sementara file yang diupload
            with open("temp_audit.xlsx", "wb") as f: 
                f.write(uploaded.getbuffer())
            
            raw_data = run_deep_audit("temp_audit.xlsx")
            
            if raw_data:
                df = pd.DataFrame(raw_data)
                
                # 1. Cek Duplikasi Internal (dalam satu file yang sama)
                df['is_internal_dup'] = df.duplicated('Hash', keep='first')
                
                # 2. Cek Duplikasi History (Database bulan sebelumnya)
                hist_check = []
                for h in df['Hash']:
                    found = cek_global_history(h)
                    hist_check.append(f"⚠️ Pernah ada di: {found[0]}" if found else "✅ FOTO BARU")
                df['History_Status'] = hist_check

                # 3. Keputusan Final Audit
                def judge(row):
                    if "⚠️" in row['History_Status']: return "❌ GUGUR (HISTORY)"
                    if row['is_internal_dup']: return "❌ GUGUR (INTERNAL)"
                    return "✅ VALID"
                
                df['Final_Result'] = df.apply(judge, axis=1)
                
                st.divider()
                st.success(f"Audit Selesai! {len(df)} foto berhasil diproses.")

                # EXCEL DOWNLOADER
                res_excel = df.drop(columns=['Img_Obj', 'Hash'])
                towrite = io.BytesIO()
                res_excel.to_excel(towrite, index=False)
                st.download_button("📥 Download Laporan Audit (.xlsx)", towrite.getvalue(), "Laporan_Audit_Duplicate.xlsx")

                # VISUALISASI
                t1, t2 = st.tabs(["📊 Hasil Tabel", "🖼️ Galeri Audit"])
                with t1:
                    st.dataframe(res_excel, use_container_width=True)
                with t2:
                    for _, r in df.iterrows():
                        color = "red" if "❌" in r['Final_Result'] else "green"
                        st.markdown(f"### :{color}[{r['Final_Result']}]")
                        st.write(f"**Cluster:** {r['Cluster']} | **Sumber:** {r['Source']} | **Lokasi:** {r['Location']}")
                        st.image(r['Img_Obj'], width=300)
                        st.divider()
            else:
                st.warning("Tidak ditemukan foto tertanam atau link Google Drive yang valid.")

    # Bersihkan file sementara
    if os.path.exists("temp_audit.xlsx"):
        os.remove("temp_audit.xlsx")
